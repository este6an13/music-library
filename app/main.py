"""
Music Library — FastAPI application.

Routes:
  GET  /         → library page (all tracks as embedded JSON)
  GET  /add      → add song page
  POST /add      → fetch Deezer metadata, download cover, store in JSON
  POST /api/fetch-track  → preview a Deezer track (returns metadata without saving)
  DELETE /track/{deezer_id}  → remove a track
  PATCH  /track/{deezer_id}/tags → update tags
"""

import json
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, Response, StreamingResponse, RedirectResponse
import csv
import io
import os
import subprocess
import openpyxl
from dotenv import load_dotenv
from pydantic import BaseModel

from .music_service import get_deezer_track, download_cover

load_dotenv()
ADMIN_MODE = os.getenv("ADMIN_MODE", "false").lower() == "true"

GCP_DATA_BUCKET_NAME = os.getenv("GCP_DATA_BUCKET_NAME", "")

app = FastAPI(title="Music Library")

def get_image_base_url() -> str:
    """Compute the base URL for serving cover images."""
    if GCP_DATA_BUCKET_NAME:
        return f"https://storage.googleapis.com/{GCP_DATA_BUCKET_NAME}/"
    return "/"


# --- Static files & templates ---
app.mount("/static", StaticFiles(directory="static"), name="static")
# Serve cover images from data/covers
DATA_DIR = Path("data")
COVERS_DIR = DATA_DIR / "covers"
LIBRARY_FILE = DATA_DIR / "library.json"

templates = Jinja2Templates(directory="templates")


# --- Data helpers ---
def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    COVERS_DIR.mkdir(parents=True, exist_ok=True)
    if not LIBRARY_FILE.exists():
        LIBRARY_FILE.write_text("[]", encoding="utf-8")


def _read_library() -> list[dict]:
    _ensure_data_dir()
    try:
        return json.loads(LIBRARY_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, FileNotFoundError):
        return []


def _write_library(tracks: list[dict]):
    _ensure_data_dir()
    LIBRARY_FILE.write_text(json.dumps(tracks, indent=2, ensure_ascii=False), encoding="utf-8")


# --- Startup ---
@app.on_event("startup")
async def startup_event():
    _ensure_data_dir()


# --- Routes ---

@app.get("/", response_class=HTMLResponse)
async def library_page(request: Request):
    """Library page — serves all tracks as embedded JSON."""
    tracks = _read_library()
    tracks_json = json.dumps(tracks, ensure_ascii=False)
    return templates.TemplateResponse(request=request, name="library.html", context={
        "request": request,
        "tracks_json": tracks_json,
        "admin_mode": ADMIN_MODE,
        "image_base_url": get_image_base_url(),
    })


@app.get("/add", response_class=HTMLResponse)
async def add_page(request: Request):
    """Add song page."""
    if not ADMIN_MODE:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(request=request, name="add.html", context={"request": request, "admin_mode": ADMIN_MODE})


class FetchTrackRequest(BaseModel):
    deezer_id: str


@app.post("/api/fetch-track")
async def fetch_track_preview(body: FetchTrackRequest):
    """Fetch track metadata from Deezer without saving. Used for preview."""
    track = await get_deezer_track(body.deezer_id)
    if not track:
        raise HTTPException(status_code=404, detail="Track not found on Deezer")
    return track


class AddTrackRequest(BaseModel):
    deezer_id: str
    tags: list[str] = []


@app.post("/api/add-track")
async def add_track(body: AddTrackRequest):
    """Fetch metadata from Deezer, download cover, and save to library."""
    if not ADMIN_MODE:
        raise HTTPException(status_code=403, detail="Not authorized")
    library = _read_library()

    # Check for duplicates
    if any(t["deezer_id"] == body.deezer_id for t in library):
        raise HTTPException(status_code=409, detail="Track already in library")

    # Fetch metadata
    track = await get_deezer_track(body.deezer_id)
    if not track:
        raise HTTPException(status_code=404, detail="Track not found on Deezer")

    # Download cover
    cover_path = await download_cover(track["cover_url"], body.deezer_id)

    # Build track entry
    entry = {
        "deezer_id": track["deezer_id"],
        "title": track["title"],
        "artist": track["artist"],
        "album": track["album"],
        "release_year": track["release_year"],
        "cover": cover_path,
        "duration": track.get("duration", 0),
        "preview_url": track.get("preview_url", ""),
        "tags": [t.strip() for t in body.tags if t.strip()],
        "added_at": datetime.utcnow().isoformat(),
    }

    library.append(entry)
    _write_library(library)

    return {"message": "success", "track": entry}


@app.delete("/api/track/{deezer_id}")
async def delete_track(deezer_id: str):
    """Remove a track from the library."""
    if not ADMIN_MODE:
        raise HTTPException(status_code=403, detail="Not authorized")
    library = _read_library()
    new_library = [t for t in library if str(t["deezer_id"]) != deezer_id]

    if len(new_library) == len(library):
        raise HTTPException(status_code=404, detail="Track not found")

    # Delete cover file if it exists
    removed = [t for t in library if str(t["deezer_id"]) == deezer_id]
    for t in removed:
        if t.get("cover"):
            cover_path = DATA_DIR / t["cover"]
            if cover_path.exists():
                cover_path.unlink()

    _write_library(new_library)
    return {"message": "deleted"}


@app.post("/api/track/{deezer_id}/refetch")
async def refetch_track(deezer_id: str):
    """Refetch track metadata and cover from Deezer."""
    if not ADMIN_MODE:
        raise HTTPException(status_code=403, detail="Not authorized")
    library = _read_library()

    # Find the track
    existing_track_idx = -1
    for i, t in enumerate(library):
        if str(t["deezer_id"]) == deezer_id:
            existing_track_idx = i
            break
            
    if existing_track_idx == -1:
        raise HTTPException(status_code=404, detail="Track not found in library")

    existing_track = library[existing_track_idx]

    # Fetch new metadata
    new_data = await get_deezer_track(deezer_id)
    if not new_data:
        raise HTTPException(status_code=404, detail="Track not found on Deezer")

    # Download new cover
    new_cover_path = await download_cover(new_data["cover_url"], deezer_id)
    
    # Merge data (keep tags and added_at)
    updated_entry = {
        "deezer_id": new_data["deezer_id"],
        "title": new_data["title"],
        "artist": new_data["artist"],
        "album": new_data["album"],
        "release_year": new_data["release_year"],
        "cover": new_cover_path or existing_track.get("cover"),
        "duration": new_data.get("duration", 0),
        "preview_url": new_data.get("preview_url", ""),
        "tags": existing_track.get("tags", []),
        "added_at": existing_track.get("added_at"),
        "isrc": new_data.get("isrc")
    }

    library[existing_track_idx] = updated_entry
    _write_library(library)

    return {"message": "success", "track": updated_entry}


class UpdateTagsRequest(BaseModel):
    tags: list[str]


@app.patch("/api/track/{deezer_id}/tags")
async def update_tags(deezer_id: str, body: UpdateTagsRequest):
    """Update tags for a track."""
    if not ADMIN_MODE:
        raise HTTPException(status_code=403, detail="Not authorized")
    library = _read_library()
    found = False
    for track in library:
        if str(track["deezer_id"]) == deezer_id:
            track["tags"] = [t.strip() for t in body.tags if t.strip()]
            found = True
            break

    if not found:
        raise HTTPException(status_code=404, detail="Track not found")

    _write_library(library)
    return {"message": "tags updated", "tags": library[0]["tags"] if library else []}

@app.post("/api/import")
async def import_library(file: UploadFile = File(...)):
    """Import tracks from CSV, JSON, or XLSX and fetch metadata with SSE progress."""
    if not ADMIN_MODE:
        raise HTTPException(status_code=403, detail="Not authorized")
    raw_content = await file.read()
    content: bytes = raw_content if isinstance(raw_content, bytes) else raw_content.encode('utf-8')
    filename = file.filename.lower() if file.filename else "unknown"
    
    tracks_to_import = []
    
    try:
        if filename.endswith(".json"):
            data = json.loads(content)
            for item in data:
                if "deezer_id" in item or "Deezer ID" in item:
                    did = str(item.get("deezer_id") or item.get("Deezer ID"))
                    tags = item.get("tags", []) if isinstance(item.get("tags"), list) else [t.strip() for t in str(item.get("tags", "")).split(",") if t.strip()]
                    tracks_to_import.append({"deezer_id": did, "tags": tags})
        
        elif filename.endswith(".csv"):
            reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
            for row in reader:
                # Support "deezer_id" or "Deezer ID"
                did_key = next((k for k in row.keys() if k and k.lower() in ["deezer_id", "deezer id"]), None)
                if did_key and row[did_key]:
                    tag_key = next((k for k in row.keys() if k and k.lower() == "tags"), None)
                    tags = [t.strip() for t in row[tag_key].split(",")] if tag_key and row[tag_key] else []
                    tracks_to_import.append({"deezer_id": str(row[did_key]), "tags": tags})
                    
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 0:
                headers = [str(cell).lower() if cell is not None else "" for cell in rows[0]]
                did_idx = -1
                tag_idx = -1
                for i, h in enumerate(headers):
                    if h in ["deezer_id", "deezer id"]: did_idx = i
                    elif h == "tags": tag_idx = i
                
                if did_idx != -1:
                    for row in rows[1:]:
                        did = str(row[did_idx]) if row[did_idx] is not None else ""
                        if did and did.lower() != "none" and did.strip():
                            tags = []
                            if tag_idx != -1 and row[tag_idx] is not None:
                                tags = [t.strip() for t in str(row[tag_idx]).split(",")]
                            tracks_to_import.append({"deezer_id": did.strip(), "tags": tags})
                            
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
            
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

    # Deduplicate against current library
    library = _read_library()
    existing_ids = {str(t["deezer_id"]) for t in library}
    
    unique_imports = []
    seen_import_ids = set()
    for t in tracks_to_import:
        if t["deezer_id"] not in existing_ids and t["deezer_id"] not in seen_import_ids:
            unique_imports.append(t)
            seen_import_ids.add(t["deezer_id"])

    async def import_generator():
        # Start SSE format
        yield f"data: {json.dumps({'status': 'start', 'total': len(unique_imports)})}\n\n"
        
        success_count = 0
        current_library = _read_library()
        
        for i, track_req in enumerate(unique_imports):
            did = track_req["deezer_id"]
            tags = track_req["tags"]
            
            try:
                # Rate limiter is called inside get_deezer_track!
                track = await get_deezer_track(did)
                if track:
                    cover_path = await download_cover(track["cover_url"], did)
                    
                    entry = {
                        "deezer_id": track["deezer_id"],
                        "title": track["title"],
                        "artist": track["artist"],
                        "album": track["album"],
                        "release_year": track["release_year"],
                        "cover": cover_path,
                        "duration": track.get("duration", 0),
                        "preview_url": track.get("preview_url", ""),
                        "tags": tags,
                        "added_at": datetime.utcnow().isoformat(),
                        "isrc": track.get("isrc")
                    }
                    current_library.append(entry)
                    _write_library(current_library)
                    success_count += 1
                    
                    yield f"data: {json.dumps({'status': 'progress', 'current': i + 1, 'total': len(unique_imports), 'track': track['title']})}\n\n"
                else:
                    yield f"data: {json.dumps({'status': 'error', 'current': i + 1, 'total': len(unique_imports), 'message': f'Deezer ID {did} not found'})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'status': 'error', 'current': i + 1, 'total': len(unique_imports), 'message': str(e)})}\n\n"
                
        yield f"data: {json.dumps({'status': 'done', 'added': success_count, 'total': len(unique_imports)})}\n\n"

    return StreamingResponse(import_generator(), media_type="text/event-stream")


@app.post("/api/sync")
async def sync_database():
    """Sync the database to Google Cloud Storage (admin only)."""
    if not ADMIN_MODE:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    data_bucket = os.getenv("GCP_DATA_BUCKET_NAME")
    if not data_bucket:
        raise HTTPException(status_code=500, detail="GCP_DATA_BUCKET_NAME not set in .env")
        
    data_dir = Path("data").resolve()
    if not data_dir.exists():
        raise HTTPException(status_code=500, detail="Local 'data' directory not found")
        
    cmd = [
        "gsutil", "-m", "rsync", "-r",
        str(data_dir),
        f"gs://{data_bucket}"
    ]
    
    meta_cmd = [
        "gsutil", "-m", "setmeta", "-h", "Cache-Control:public, max-age=31536000, immutable",
        f"gs://{data_bucket}/covers/*.jpg"
    ]
    
    try:
        is_windows = os.name == 'nt'
        result = subprocess.run(cmd, check=True, shell=is_windows, capture_output=True, text=True)
        # Attempt to set Cache-Control on all .jpg items in covers bucket
        subprocess.run(meta_cmd, shell=is_windows, capture_output=True, text=True)
        return {"message": "Sync completed successfully\n" + result.stdout[:500]}
    except subprocess.CalledProcessError as e:
        raise HTTPException(status_code=500, detail=f"Sync failed with exit code {e.returncode}: {e.stderr}")
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="Error: 'gsutil' command not found. Please ensure Google Cloud CLI is installed.")


@app.get("/api/export/{fmt}")
async def export_library(fmt: str):
    """Export the music library to CSV, JSON, or XLSX."""
    library = _read_library()
    
    timestamp = datetime.now().strftime("%y%m%d-%H%M%S")
    filename = f"music-library-data-{timestamp}.{fmt}"

    if fmt == "json":
        # Stripping out internal paths like 'cover' before exporting
        export_data = [{k: v for k, v in track.items() if k != 'cover'} for track in library]
        content = json.dumps(export_data, indent=2, ensure_ascii=False)
        return Response(
            content=content,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    # Prepare flat data for CSV and XLSX
    headers = ["Title", "Artist", "Album", "Year", "Duration (s)", "Tags", "Deezer ID", "ISRC"]
    rows = []
    for track in library:
        rows.append([
            track.get("title", ""),
            track.get("artist", ""),
            track.get("album", ""),
            track.get("release_year", ""),
            track.get("duration", 0),
            ", ".join(track.get("tags", [])),
            track.get("deezer_id", ""),
            track.get("isrc", "")
        ])

    if fmt == "csv":
        output = io.BytesIO()
        # Write UTF-8 BOM so Excel/spreadsheet apps detect the encoding correctly
        wrapper = io.TextIOWrapper(output, encoding="utf-8-sig", newline="")
        writer = csv.writer(wrapper)
        writer.writerow(headers)
        writer.writerows(rows)
        wrapper.flush()
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    if fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Music Library"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        
        output = io.BytesIO()
        wb.save(output)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    raise HTTPException(status_code=400, detail="Invalid format. Supported: json, csv, xlsx")
